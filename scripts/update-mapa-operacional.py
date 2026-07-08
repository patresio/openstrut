#!/usr/bin/env python3
"""Update mapa_operacional.xlsx with OpenTrust runtime entries."""

import openpyxl
from copy import copy
from pathlib import Path

XLSX = Path(__file__).parent.parent / 'mapa_operacional.xlsx'
wb = openpyxl.load_workbook(XLSX)

# ── Data ──────────────────────────────────────────────────────────────────────

OPENTRUST_AGENTS = [
    # Leaders
    ('OTLD01', 'trust-lead', 'Coordinates cross-team communication, decision logging, meeting facilitation, and process health',
     'Phase 1-2 (Explore/Propose), cross-team coordination, workflow governance',
     'Coordinate OpenTrust workflow, decisions, handoffs, and process health',
     'opentrust-task-contract; opentrust-observability; opentrust-review',
     '', '', '', '', '', '', 'global/agents/trust-lead.md',
     'opentrust-task-contract; opentrust-observability', 'coordination-facilitator; meeting-scribe; decision-logger',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD02', 'product-lead', 'Product strategy, discovery, requirements, acceptance criteria, and story slicing',
     'Phase 1-2 (Explore/Propose), requirements gathering, scope definition',
     'Turn vague requests into actionable, testable scope',
     'opentrust-task-contract; opentrust-spec-change',
     '', '', '', '', '', '', 'global/agents/product-lead.md',
     'opentrust-task-contract; opentrust-spec-change', 'product-discovery; requirements-analyzer; story-slicer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD03', 'architecture-lead', 'Structural decisions, domain modeling, API/database contracts, distributed systems, and ADRs',
     'Phase 2-3 (Propose/Apply), architecture decisions, trade-off analysis',
     'Guide structural decisions and trade-offs',
     'opentrust-spec-change; opentrust-task-contract',
     '', '', '', '', '', '', 'global/agents/architecture-lead.md',
     'opentrust-spec-change; opentrust-task-contract', 'architecture-decision-designer; domain-modeler; api-database-designer; distributed-systems-reviewer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD04', 'engineering-lead', 'Implementation, refactoring, performance, security, and privacy leadership',
     'Phase 3 (Apply), implementation, code changes within approved scope',
     'Coordinate implementation work inside approved scope',
     'opentrust-tdd; opentrust-task-contract',
     '', '', '', '', '', '', 'global/agents/engineering-lead.md',
     'opentrust-tdd; opentrust-task-contract', 'feature-implementer; code-refactoring-specialist; performance-engineer; security-reviewer; privacy-reviewer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD05', 'quality-lead', 'Test strategy, TDD, integration tests, end-to-end tests, and quality gates',
     'Phase 3-4 (Apply/Review), test planning, quality verification',
     'Define and verify quality strategy',
     'opentrust-tdd; opentrust-review',
     '', '', '', '', '', '', 'global/agents/quality-lead.md',
     'opentrust-tdd; opentrust-review', 'tdd-engineer; integration-tester; testing-strategy-designer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD06', 'review-lead', 'Independent review, compliance, UX/accessibility review, and delivery gating',
     'Phase 4 (Review), pre-delivery quality gate, compliance check',
     'Gate changes through independent review',
     'opentrust-review',
     '', '', '', '', '', '', 'global/agents/review-lead.md',
     'opentrust-review', 'code-reviewer; compliance-auditor; ux-accessibility-reviewer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD07', 'devops-lead', 'CI/CD, infrastructure, observability, and incident response leadership',
     'Any phase, infrastructure/ops decisions, incident triage',
     'Coordinate delivery infrastructure and operational reliability',
     'opentrust-observability; opentrust-task-contract',
     '', '', '', '', '', '', 'global/agents/devops-lead.md',
     'opentrust-observability; opentrust-task-contract', 'ci-cd-infrastructure-engineer; observability-designer; incident-triage-specialist',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD08', 'delivery-lead', 'Release management, versioning, changelog, and deployment coordination',
     'Phase 5 (Ship), post-review release preparation',
     'Prepare approved work for release',
     'opentrust-delivery; opentrust-observability',
     '', '', '', '', '', '', 'global/agents/delivery-lead.md',
     'opentrust-delivery; opentrust-observability', 'release-manager; changelog-writer',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    ('OTLD09', 'knowledge-lead', 'Context retrieval, reference library management, documentation generation, and skill creation',
     'All phases, when retrieval or reference material is needed',
     'Provide selector-based synthesis and reference governance. Only team that calls the retrieval provider directly.',
     'opentrust-reference-research; opentrust-task-contract',
     '', '', '', '', '', '', 'global/agents/knowledge-lead.md',
     'opentrust-reference-research; opentrust-task-contract', 'context-historian; reference-librarian; documentation-skill-creator',
     '', 'primary', 'opentrust', 'opentrust-runtime', '', '', '',
     'yes', 'OpenTrust team lead'),
    # Subagents — Trust Coordination
    ('OTAG01', 'coordination-facilitator', 'Facilitates Trust Coordination handoffs, meetings, and cross-team alignment',
     'Phase 1-2, cross-team coordination, blocker resolution',
     'Facilitate cross-team coordination, handoffs, and blocker alignment',
     'opentrust-task-contract', '', '', '', '', '', '', 'global/agents/coordination-facilitator.md',
     'opentrust-task-contract', 'trust-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG02', 'meeting-scribe', 'Captures meeting notes, action items, and handoff evidence for Trust Coordination',
     'After meetings, handoffs, decision points',
     'Capture concise meeting notes, action items, open questions, and evidence',
     'opentrust-observability', '', '', '', '', '', '', 'global/agents/meeting-scribe.md',
     'opentrust-observability', 'trust-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG03', 'decision-logger', 'Records decisions, rationale, and unresolved risks for Trust Coordination',
     'After decisions, trade-off evaluations, blocked tasks',
     'Record decisions, rationale, alternatives rejected, risks, and follow-up needs',
     'opentrust-observability', '', '', '', '', '', '', 'global/agents/decision-logger.md',
     'opentrust-observability', 'trust-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Product / Discovery
    ('OTAG04', 'product-discovery', 'Explores product problems, users, outcomes, and discovery risks',
     'Phase 1 (Explore), before shaping any product work',
     'Clarify product problem, users, outcomes, assumptions, risks, and discovery questions',
     'opentrust-task-contract', '', '', '', '', '', '', 'global/agents/product-discovery.md',
     'opentrust-task-contract', 'product-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG05', 'requirements-analyzer', 'Analyzes requirements, constraints, acceptance criteria, and gaps',
     'Phase 2 (Propose), requirements refinement',
     'Analyze requirements, constraints, dependencies, acceptance criteria, and missing details',
     'opentrust-task-contract', '', '', '', '', '', '', 'global/agents/requirements-analyzer.md',
     'opentrust-task-contract', 'product-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG06', 'story-slicer', 'Slices product work into small testable stories with explicit exclusions',
     'Phase 2 (Propose), after requirements are clear',
     'Slice scoped work into small testable stories, edges, and explicit out-of-scope items',
     'opentrust-task-contract; opentrust-tdd', '', '', '', '', '', '', 'global/agents/story-slicer.md',
     'opentrust-task-contract; opentrust-tdd', 'product-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Architecture
    ('OTAG07', 'architecture-decision-designer', 'Compares architecture options, trade-offs, and ADR-ready decisions',
     'Phase 2-3 (Propose/Apply), architecture decisions',
     'Compare architecture options, constraints, trade-offs, and ADR-ready decision points',
     'opentrust-spec-change', '', '', '', '', '', '', 'global/agents/architecture-decision-designer.md',
     'opentrust-spec-change', 'architecture-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG08', 'domain-modeler', 'Shapes domain models, concepts, boundaries, and core language',
     'Phase 2-3 (Propose/Apply), domain modeling',
     'Define domain concepts, boundaries, relationships, invariants, and shared language',
     'opentrust-spec-change', '', '', '', '', '', '', 'global/agents/domain-modeler.md',
     'opentrust-spec-change', 'architecture-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG09', 'api-database-designer', 'Designs APIs, schemas, data contracts, and evolution paths',
     'Phase 3 (Apply), API/database design',
     'Design APIs, schemas, data models, contracts, evolution, and migration paths',
     'opentrust-spec-change', '', '', '', '', '', '', 'global/agents/api-database-designer.md',
     'opentrust-spec-change', 'architecture-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG10', 'distributed-systems-reviewer', 'Reviews distributed-systems patterns, consistency, trade-offs, and failure modes',
     'Phase 4 (Review), distributed systems audit',
     'Review consistency, concurrency, failures, retries, eventual consistency, and trade-offs',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/distributed-systems-reviewer.md',
     'opentrust-review', 'architecture-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Engineering
    ('OTAG11', 'feature-implementer', 'Build approved feature changes in small validated increments',
     'Phase 3 (Apply), feature implementation',
     'Implements approved changes. Keeps diff small. Runs focused validation.',
     'opentrust-tdd', '', '', '', '', '', '', 'global/agents/feature-implementer.md',
     'opentrust-tdd', 'engineering-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG12', 'code-refactoring-specialist', 'Refactor code incrementally while preserving approved behavior',
     'Phase 3 (Apply), refactoring tasks',
     'Improves structure with behavior preserved. Prefers smallest safe change.',
     'opentrust-tdd', '', '', '', '', '', '', 'global/agents/code-refactoring-specialist.md',
     'opentrust-tdd', 'engineering-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG13', 'performance-engineer', 'Find and reduce performance bottlenecks with measured evidence',
     'Phase 3-4 (Apply/Review), performance optimization',
     'Profiles hot paths. Suggests smallest fix with evidence.',
     'opentrust-tdd; opentrust-review', '', '', '', '', '', '', 'global/agents/performance-engineer.md',
     'opentrust-tdd; opentrust-review', 'engineering-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG14', 'security-reviewer', 'Review implementation for security risks and minimum effective controls',
     'Phase 4 (Review), security audit',
     'Checks auth, secrets, trust boundaries, unsafe defaults, dependency risk',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/security-reviewer.md',
     'opentrust-review', 'engineering-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG15', 'privacy-reviewer', 'Review privacy impact, data handling, retention, and exposure risks',
     'Phase 4 (Review), privacy audit',
     'Checks data minimization, consent, retention, sharing, logging exposure',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/privacy-reviewer.md',
     'opentrust-review', 'engineering-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Testing / Quality
    ('OTAG16', 'tdd-engineer', 'Drive red-green-refactor for smallest useful automated tests',
     'Phase 3 (Apply), test-first for behavioral changes',
     'Defines failing test first. Confirms red and green evidence.',
     'opentrust-tdd', '', '', '', '', '', '', 'global/agents/tdd-engineer.md',
     'opentrust-tdd', 'quality-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG17', 'integration-tester', 'Verify integrated behavior across boundaries with focused evidence',
     'Phase 4 (Review), integration test verification',
     'Checks behavior across modules, contracts, fixtures, and regressions',
     'opentrust-tdd; opentrust-review', '', '', '', '', '', '', 'global/agents/integration-tester.md',
     'opentrust-tdd; opentrust-review', 'quality-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG18', 'testing-strategy-designer', 'Design lean test strategy matched to risk, scope, and feedback speed',
     'Phase 2-3 (Propose/Apply), test strategy definition',
     'Chooses test levels, coverage targets, and smallest useful checks',
     'opentrust-task-contract; opentrust-tdd', '', '', '', '', '', '', 'global/agents/testing-strategy-designer.md',
     'opentrust-task-contract; opentrust-tdd', 'quality-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Review / Governance (includes preserved legacy)
    ('OTAG19', 'code-reviewer', 'Read-only independent review of implementation diff, tests, scope, security, contracts, migrations, dependencies, and regressions before delivery',
     'Phase 4 (Review), pre-delivery code review',
     'Read-only code review. Checks correctness, scope compliance, security, privacy, contracts, dependencies, legacy preservation, error paths, docs.',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/code-reviewer.md',
     'opentrust-review', 'review-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', 'Preserved legacy agent'),
    ('OTAG20', 'compliance-auditor', 'Auditar conformidade — dependencias, licencas, vulnerabilidades, supply chain, regulamentacao e politicas',
     'Phase 4 (Review), compliance/security audit',
     'Dependency audit, license checks, vulnerability scanning, supply chain verification, regulatory compliance',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/compliance-auditor.md',
     'opentrust-review', 'review-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', 'Preserved legacy agent'),
    ('OTAG21', 'ux-accessibility-reviewer', 'Review user experience and accessibility issues before delivery gates',
     'Phase 4 (Review), UX/accessibility audit',
     'Checks flows, clarity, keyboard access, semantics, and blocking UX risks',
     'opentrust-review', '', '', '', '', '', '', 'global/agents/ux-accessibility-reviewer.md',
     'opentrust-review', 'review-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — DevOps / SRE
    ('OTAG22', 'ci-cd-infrastructure-engineer', 'Designs CI/CD flow, pipeline guards, and infrastructure delivery checks',
     'Phase 3-5 (Apply to Ship), CI/CD/infra decisions',
     'Shapes CI/CD and infra rollout plan. Checks pipeline gates, rollback path, and deploy safety.',
     'opentrust-observability; opentrust-delivery', '', '', '', '', '', '', 'global/agents/ci-cd-infrastructure-engineer.md',
     'opentrust-observability; opentrust-delivery', 'devops-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG23', 'observability-designer', 'Designs logs, metrics, traces, alerts, and debug paths',
     'Phase 3-5 (Apply to Ship), observability definition',
     'Defines observability plan for logs, metrics, traces, and alerts.',
     'opentrust-observability', '', '', '', '', '', '', 'global/agents/observability-designer.md',
     'opentrust-observability', 'devops-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG24', 'incident-triage-specialist', 'Triage incidents with smallest safe containment and recovery steps',
     'Incident phase, production/deployment issues',
     'Triage incidents, narrow blast radius, and propose recovery order.',
     'opentrust-observability', '', '', '', '', '', '', 'global/agents/incident-triage-specialist.md',
     'opentrust-observability', 'devops-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Delivery / Release (includes preserved legacy)
    ('OTAG25', 'release-manager', 'Gerenciar releases, changelog, versionamento semantico, depreciacao e notas de release',
     'Phase 5 (Ship), release planning and execution',
     'Plan, document and coordinate releases - semver, changelog, deprecation, release notes',
     'opentrust-delivery; opentrust-observability', '', '', '', '', '', '', 'global/agents/release-manager.md',
     'opentrust-delivery; opentrust-observability', 'delivery-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', 'Preserved legacy agent'),
    ('OTAG26', 'changelog-writer', 'Produces compact changelog entries from approved diffs and release facts',
     'Phase 5 (Ship), changelog generation',
     'Drafts changelog notes from approved scope, diff facts, and validation evidence.',
     'opentrust-delivery', '', '', '', '', '', '', 'global/agents/changelog-writer.md',
     'opentrust-delivery', 'delivery-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    # Subagents — Knowledge / Context
    ('OTAG27', 'context-historian', 'Tracks context lineage, decisions, and selector history across tasks',
     'All phases, retrieval coherence tasks',
     'Tracks context decisions, selector use, and knowledge continuity across tasks.',
     'opentrust-reference-research; opentrust-observability', '', '', '', '', '', '', 'global/agents/context-historian.md',
     'opentrust-reference-research; opentrust-observability', 'knowledge-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG28', 'reference-librarian', 'Curates selector mappings, bundles, and reference profile consistency',
     'All phases, selector/bundle curation',
     'Curates selector sets, bundle fit, and reference profile consistency.',
     'opentrust-reference-research', '', '', '', '', '', '', 'global/agents/reference-librarian.md',
     'opentrust-reference-research', 'knowledge-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
    ('OTAG29', 'documentation-skill-creator', 'Shapes documentation and skill drafts around approved selectors and source policy',
     'All phases, doc/skill creation',
     'Prepares documentation and skill creation guidance from approved selectors and reference policy.',
     'opentrust-reference-research', '', '', '', '', '', '', 'global/agents/documentation-skill-creator.md',
     'opentrust-reference-research', 'knowledge-lead', '', 'subagent', 'opentrust', 'opentrust-runtime', '', '', '', 'yes', ''),
]

OPENTRUST_SKILLS = [
    ('OTSK01', 'opentrust-task-contract', 'Create or refine task contracts using TASK_CONTRACT.md template including Retrieval Context selectors',
     'When a task needs formal scope, acceptance criteria, retrieval requirements, and delivery rules before implementation',
     'task request; scope description; user intent', 'task contract with acceptance criteria, retrieval selectors, teams involved',
     '', '', '', '', 'global/skills/opentrust-task-contract/SKILL.md',
     'Contract must reference approved selectors; no raw chunks in output',
     'Must include Retrieval Context section', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD01; OTLD02; OTLD03; OTLD04; OTLD08', 'OTSK02; OTSK03; OTSK05', 'yes', 'opentrust; workflow'),
    ('OTSK02', 'opentrust-tdd', 'Enforce test-first workflow when behavior or rules change — define failing test, implement, validate',
     'When changing executable behavior (bugfix, new feature, refactoring of behavior)',
     'behavior description; acceptance criteria; existing test files', 'RED evidence; GREEN evidence; refactored code; test results',
     '', '', '', '', 'global/skills/opentrust-tdd/SKILL.md',
     'Must show RED before implementation; GREEN must verify behavior',
     'Not for documentation-only changes', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD04; OTLD05; OTAG11; OTAG16', 'OTSK01; OTSK03', 'yes', 'opentrust; workflow; tdd'),
    ('OTSK03', 'opentrust-spec-change', 'Guide structured spec and design changes using Explore -> Propose before Apply',
     'When a feature or change requires formal specification, business rule analysis, impact assessment, or an approved plan before code',
     'change request; context; affected boundaries', 'spec document; task plan; approved plan with AC',
     '', '', '', '', 'global/skills/opentrust-spec-change/SKILL.md',
     'Spec must include AC, scope, risks, and retrieval selectors',
     'Not for trivial or fully-scoped tasks', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'sdd; OTLD02; OTLD03; OTLD04', 'OTSK01; OTSK02', 'yes', 'opentrust; workflow; spec'),
    ('OTSK04', 'opentrust-review', 'Review diff, tests, security, architecture, and docs — approve or block with evidence',
     'When a diff is ready for independent review before delivery',
     'diff; implementation plan; test results; AC', 'review report with findings, blocking issues, or approval',
     '', '', '', '', 'global/skills/opentrust-review/SKILL.md',
     'Must verify AC, scope, security, edge cases, tests, and docs',
     'Not for trivial formatting-only changes', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD06; OTAG19', 'OTSK01; OTSK02', 'yes', 'opentrust; workflow; review'),
    ('OTSK05', 'opentrust-delivery', 'Prepare commit, push, and pull request using Conventional Commits in English. Avoid unrelated files.',
     'When approved work is ready for version control and delivery',
     'approved scope; diff; validation evidence', 'commit; push; PR (when requested)',
     '', '', '', '', 'global/skills/opentrust-delivery/SKILL.md',
     'Only deliver approved scope; never include unrelated files',
     'Not for unapproved changes', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD08; OTAG25', 'OTSK04; OTSK06', 'yes', 'opentrust; workflow; delivery'),
    ('OTSK06', 'opentrust-observability', 'Require execution reports, validation evidence, and operational notes. Do not implement external telemetry yet.',
     'When work requires traceability, decision logging, or deployment evidence',
     'phase; task; decisions made', 'execution report with validation evidence and operational notes',
     '', '', '', '', 'global/skills/opentrust-observability/SKILL.md',
     'Must log decisions, evidence, blockers, and next actions',
     'Not for external telemetry or monitoring infrastructure', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD01; OTLD07; OTAG23', 'OTSK01; OTSK04', 'yes', 'opentrust; workflow; observability'),
    ('OTSK07', 'opentrust-reference-research', 'Use Operational Retrieval Map selectors. Request synthesis only. No raw chunks in output or commits.',
     'When domain knowledge, architecture patterns, or reference material is needed during any workflow phase',
     'selector query (CTX/SK/B/DOC); task contract', 'synthesized response with source IDs; no raw chunks',
     '', '', '', '', 'global/skills/opentrust-reference-research/SKILL.md',
     'Must cite source IDs; no raw chunks in versioned output',
     'Not for trivial or fully-scoped tasks', 'active', 'opentrust', 'opentrust-runtime', '', '', '',
     'OTLD09; OTAG28; OTAG27', 'OTSK01; OTSK03', 'yes', 'opentrust; workflow; retrieval'),
]

OPENTRUST_COMMANDS = [
    ('OTC01', 'ot-explore', 'Perform read-only exploration: inspect current state, identify scope, files, risks, and needed Retrieval Context selectors',
     'global/commands/ot-explore.md', 'active', 'Phase 1 — before any work begins, to understand current state and scope'),
    ('OTC02', 'ot-propose', 'Produce task contract with acceptance criteria, test plan, and Retrieval Context selectors',
     'global/commands/ot-propose.md', 'active', 'Phase 2 — after Explore, to formalize scope and get approval'),
    ('OTC03', 'ot-apply', 'Implement only an approved task contract with TDD-First for behavioral changes',
     'global/commands/ot-apply.md', 'active', 'Phase 3 — after proposal approval, to implement one microincrement at a time'),
    ('OTC04', 'ot-review', 'Independent review of implementation diff, validation evidence, tests, security, architecture, and docs',
     'global/commands/ot-review.md', 'active', 'Phase 4 — after Apply, before delivery, for independent quality gate'),
    ('OTC05', 'ot-ship', 'Prepare commit, PR, and release artifacts only when explicitly requested. Uses Conventional Commits in English.',
     'global/commands/ot-ship.md', 'active', 'Phase 5 — after Review approval, for final delivery'),
    ('OTC06', 'ot-status', 'Summarize current state: branch, diff, active task, pending validation, and next step',
     'global/commands/ot-status.md', 'active', 'Any phase — for status reporting and context resumption'),
    ('OTC07', 'ot-incident', 'Diagnose incident or failure, collect evidence, propose fix, validate resolution',
     'global/commands/ot-incident.md', 'active', 'When a runtime, test, build, deployment, or configuration incident occurs'),
]

OPENTRUST_WORKFLOWS = [
    ('OTF01', 'opentrust-explore', 'Read-only exploration: inspect current state, identify scope, risks, and needed selectors',
     'read-only', 'Inspect > identify scope > assess risk > propose selectors',
     'OTSK01; OTSK07', 'OTLD01; OTLD02; OTLD03; OTLD09', 'global/commands/ot-explore.md', 'active', 'Phase 1 — before any work begins'),
    ('OTF02', 'opentrust-propose', 'Produce approved task contract with acceptance criteria, test plan, and retrieval selectors',
     'read-only', 'Write proposal > compare alternatives > define AC > list selectors > approval gate',
     'OTSK01; OTSK03', 'OTLD01; OTLD02; OTLD03; OTLD09', 'global/commands/ot-propose.md', 'active', 'Phase 2 — after Explore, before implementation'),
    ('OTF03', 'opentrust-apply', 'Implement approved task contract with TDD-First for behavioral changes',
     'sequential', 'Create task plan > TDD RED > implement > TDD GREEN > validate > update plan',
     'OTSK02; OTSK03', 'OTLD04; OTLD05; OTAG11; OTAG12', 'global/commands/ot-apply.md', 'active', 'Phase 3 — implementation microincrements'),
    ('OTF04', 'opentrust-review', 'Independent review of diff, tests, security, architecture, and docs',
     'read-only', 'Read diff > verify AC > run tests > check security > review docs > approve/block',
     'OTSK04', 'OTLD06; OTAG19; OTAG20; OTAG21', 'global/commands/ot-review.md', 'active', 'Phase 4 — pre-delivery quality gate'),
    ('OTF05', 'opentrust-ship', 'Prepare commit, push, and PR using Conventional Commits',
     'delivery', 'Archive > commit > push > PR > tag > release',
     'OTSK05; OTSK06', 'OTLD08; OTAG25; OTAG26', 'global/commands/ot-ship.md', 'active', 'Phase 5 — final delivery'),
    ('OTF06', 'opentrust-incident', 'Triage incidents with smallest safe containment and recovery steps',
     'sequential', 'Diagnose > contain > recover > validate > report',
     'OTSK06', 'OTLD07; OTAG24', 'global/commands/ot-incident.md', 'active', 'Incident response — any phase'),
]

OPENTRUST_BUNDLES = [
    ('B25', 'opentrust-workflow-core', 'Core workflow contract, TDD, and spec-change for OpenTrust',
     'OTSK01; OTSK02; OTSK03', 'OTLD01; OTLD02; OTLD03; OTLD04; OTLD05; OTLD08',
     'Explore > Propose > Apply phases. Use when: starting a new task, defining scope, planning implementation.'),
    ('B26', 'opentrust-delivery-core', 'Review, delivery, and observability for OpenTrust',
     'OTSK04; OTSK05; OTSK06', 'OTLD06; OTLD07; OTLD08; OTLD09',
     'Review > Ship phases. Use when: reviewing a diff, preparing release, logging decisions.'),
    ('B27', 'opentrust-retrieval-core', 'Reference research and retrieval governance for OpenTrust',
     'OTSK07', 'OTLD09; OTAG27; OTAG28; OTAG29',
     'Any phase. Use when: domain knowledge is needed, task contract specifies selectors.'),
]

OPENTRUST_PROJECTS = [
    ('PRJ13', 'OpenStrut / OpenTrust', 'Versioned engineering harness and runtime for OpenCode with OpenTrust topology',
     '', '', 'OTSK01; OTSK02; OTSK03; OTSK04; OTSK05; OTSK06; OTSK07',
     'OTLD01; OTLD02; OTLD03; OTLD04; OTLD05; OTLD06; OTLD07; OTLD08; OTLD09',
     '', 'How to install? How to update agents? How to create a task contract? How to run a workflow?',
     'Installed runtime with 68 artifacts, 38 agents, 7 commands, 7 skills'),
]

OPENTRUST_DOCS = [
    ('DOC17', 'OpenTrust Team Topology', 'opentrust/docs/TEAM_TOPOLOGY.md',
     '9-team topology, workflow phases, and agent interaction model for OpenTrust',
     'OTSK01; OTSK07', 'OTLD01; OTLD09', 'OpenStrut',
     'When understanding team roles, workflow phases, or agent coordination',
     'Not for runtime or tool-specific questions',
     'Consulte a documentação de topologia OpenTrust para entender {dúvida}'),
    ('DOC18', 'OpenTrust Workflow', 'opentrust/docs/WORKFLOW.md',
     'Workflow phases, gates, retrieval rules, and transition checklist for OpenTrust',
     'OTSK01; OTSK03; OTSK04', 'OTLD01; OTLD04; OTLD06', 'OpenStrut',
     'When planning or executing a workflow phase transition',
     'Not for agent-specific prompts or skill details',
     'Consulte o workflow OpenTrust para {fase} as regras de transição'),
    ('DOC19', 'OpenTrust Task Contract', 'opentrust/docs/TASK_CONTRACT.md',
     'Task contract template with retrieval context, teams, and Definition of Done',
     'OTSK01; OTSK03', 'OTLD01; OTLD02; OTLD03; OTLD04; OTLD08', 'OpenStrut',
     'When formalizing scope, AC, retrieval requirements, and delivery rules',
     'Not for informal or trivial tasks without retrieval needs',
     'Consulte o template de task contract para {tarefa}'),
    ('DOC20', 'OpenTrust Operational Retrieval Map', 'opentrust/docs/OPERATIONAL_RETRIEVAL_MAP.md',
     'Three-layer retrieval architecture: selectors, reference map, and provider policy',
     'OTSK07', 'OTLD09; OTAG27; OTAG28', 'OpenStrut',
     'When understanding selector system, reference profiles, or retrieval policy',
     'Not for actual retrieval queries (use provider directly)',
     'Consulte o mapa de recuperação operacional para {seletor}'),
]

OPENTRUST_PROMPTS = [
    ('P25', 'OpenTrust — iniciar workflow', 'Iniciar um novo workflow OpenTrust: explorar, propor, implementar',
     'OTLD01', 'OTSK01', 'B25',
     'TEAM_TOPOLOGY.md; WORKFLOW.md; TASK_CONTRACT.md',
     'Objetivo: {objetivo}. Contexto: {contexto}. Siga o workflow OpenTrust: 1) Explore o estado atual, 2) Proponha um task contract com acceptance criteria e selectores de retrieval, 3) Obtenha aprovação antes de implementar.',
     'Task contract aprovado, plano de execução, próximos passos'),
    ('P26', 'OpenTrust — revisar diff', 'Revisar um diff de implementação antes do delivery',
     'OTLD06', 'OTSK04', 'B26',
     'diff da implementação; task contract aprovado; resultados de teste',
     'Revisão do diff #{pr}. Verifique: 1) Acceptance criteria atendidos, 2) Cobertura de testes, 3) Segurança e privacidade, 4) Contratos e migrações, 5) Documentação. Aprova ou bloqueie com evidência.',
     'Relatório de revisão com achados, aprovação ou bloqueios'),
    ('P27', 'OpenTrust — fazer release', 'Preparar commit, push e PR para entrega',
     'OTLD08', 'OTSK05', 'B26',
     'diff final; relatório de revisão; resultados de validação',
     'Release {versao}: 1) Arquive documentação, 2) Commit com Conventional Commit, 3) Push, 4) Crie PR, 5) Tag e release. Use apenas arquivos aprovados.',
     'Commit, push, PR criado, release publicado'),
    ('P28', 'OpenTrust — consultar referência', 'Pesquisar material de referência usando selectores OpenTrust',
     'OTLD09', 'OTSK07', 'B27',
     'OPERATIONAL_RETRIEVAL_MAP.md; selectores CTX/SK/B/DOC',
     'Selectores: {selectores}. Contexto: {contexto}. Consulte o Operational Retrieval Map, peça síntese ao provider de retrieval. Retorne apenas síntese, sem chunks brutos. Inclua source IDs.',
     'Síntese com source IDs, sem chunks brutos'),
    ('P29', 'OpenTrust — investigar incidente', 'Triar e recuperar de incidente com evidência',
     'OTLD07', 'OTSK06', '',
     'logs; métricas; mudanças recentes; task contract ativo',
     'Incidente: {descricao}. 1) Diagnostique coletando evidência, 2) Contenha o dano, 3) Proponha recuperação, 4) Valide a resolução, 5) Reporte.',
     'Diagnóstico, plano de contenção, recuperação validada, relatório'),
    ('P30', 'OpenTrust — verificar status', 'Resumir estado atual do workflow e próximos passos',
     'OTLD01', 'OTSK06', '',
     'git status; diff; task plan ativo; últimas validações',
     'Resuma: branch, diff, tarefa ativa, fase do workflow, validações pendentes, próximo passo.',
     'Resumo de estado com próximo passo claro'),
]

# ── Helper ─────────────────────────────────────────────────────────────────────

def append_rows(ws, rows):
    """Append rows to worksheet preserving formatting of last row."""
    last_row = ws.max_row
    # Get format from header row as default
    header_format = None
    if ws.max_row >= 1:
        for cell in ws[1]:
            header_format = cell
            break

    for row_idx, row_data in enumerate(rows, start=last_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if header_format and ws.max_row == 1:
                cell.font = copy(header_format.font)
                cell.alignment = copy(header_format.alignment)


def ensure_sheet(wb, name, headers):
    """Get or create a worksheet with headers."""
    if name in wb.sheetnames:
        ws = wb[name]
        # Check if headers match
        existing = [cell.value for cell in ws[1]]
        if existing == headers:
            return ws
        print(f"  WARNING: {name} headers mismatch, appending anyway")
        return ws
    ws = wb.create_sheet(name)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    return ws


# ── Update sheets ─────────────────────────────────────────────────────────────

# 04_AGENTS
print("Updating 04_AGENTS...")
ws = wb['04_AGENTS']
expected_agents = ['agent_id', 'agent_name', 'role', 'when_use', 'responsibilities', 'skills', 'contexts',
                   'primary_sources', 'workflow', 'output_contract', 'prompt_skeleton', 'limits', 'file_path',
                   'status', 'runtime_scope', 'source_type', 'barsa_collection', 'barsa_contexts',
                   'barsa_bundles', 'primary_skills', 'support_skills', 'cowork_agents', 'workflow_mode',
                   'install_global', 'dedupe_note']
existing = [cell.value for cell in ws[1]]
if existing == expected_agents:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_agents}, got={existing}")

# Build agent rows from data
agent_rows = []
for a in OPENTRUST_AGENTS:
    agent_rows.append(a)
append_rows(ws, agent_rows)
print(f"  Added {len(agent_rows)} OpenTrust agents (total: {ws.max_row - 1})")

# 03_SKILLS
print("Updating 03_SKILLS...")
ws = wb['03_SKILLS']
expected_skills = ['skill_id', 'skill_name', 'objective', 'trigger', 'inputs', 'outputs', 'source_contexts',
                   'steps', 'file_path', 'tests', 'quality', 'limits', 'status', 'runtime_scope', 'source_type',
                   'barsa_collection', 'barsa_contexts', 'barsa_bundles', 'usable_by_agents', 'cowork_with',
                   'install_global', 'workflow_tags', 'dedupe_note']
existing = [cell.value for cell in ws[1]]
if existing == expected_skills:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_skills}, got={existing}")

skill_rows = []
for s in OPENTRUST_SKILLS:
    skill_rows.append(s)
append_rows(ws, skill_rows)
print(f"  Added {len(skill_rows)} OpenTrust skills (total: {ws.max_row - 1})")

# 11_COMMANDS
print("Updating 11_COMMANDS...")
ws = wb['11_COMMANDS']
expected_cmds = ['command_id', 'command_name', 'description', 'file_path', 'status', 'when_use']
existing = [cell.value for cell in ws[1]]
if existing == expected_cmds:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_cmds}, got={existing}")

cmd_rows = []
for c in OPENTRUST_COMMANDS:
    cmd_rows.append(c)
append_rows(ws, cmd_rows)
print(f"  Added {len(cmd_rows)} OpenTrust commands (total: {ws.max_row - 1})")

# 09_WORKFLOWS
print("Updating 09_WORKFLOWS...")
ws = wb['09_WORKFLOWS']
expected_wf = ['workflow_id', 'workflow_name', 'description', 'mode', 'steps', 'primary_skills',
               'primary_agents', 'file_path', 'status', 'when_use']
existing = [cell.value for cell in ws[1]]
if existing == expected_wf:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_wf}, got={existing}")

wf_rows = []
for w in OPENTRUST_WORKFLOWS:
    wf_rows.append(w)
append_rows(ws, wf_rows)
print(f"  Added {len(wf_rows)} OpenTrust workflows (total: {ws.max_row - 1})")

# 05_BUNDLES
print("Updating 05_BUNDLES...")
ws = wb['05_BUNDLES']
expected_bundles = ['bundle_id', 'bundle_name', 'problem_target', 'sources', 'order', 'skills', 'agents', 'prompt']
existing = [cell.value for cell in ws[1]]
if existing == expected_bundles:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_bundles}, got={existing}")

bundle_rows = []
for b in OPENTRUST_BUNDLES:
    bundle_rows.append(b)
append_rows(ws, bundle_rows)
print(f"  Added {len(bundle_rows)} OpenTrust bundles (total: {ws.max_row - 1})")

# 06_PROJETOS
print("Updating 06_PROJETOS...")
ws = wb['06_PROJETOS']
expected_proj = ['project_id', 'project_name', 'objective', 'contexts', 'bundles', 'skills', 'agents',
                 'source_priorities', 'typical_questions', 'expected_outputs']
existing = [cell.value for cell in ws[1]]
if existing == expected_proj:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_proj}, got={existing}")

proj_rows = []
for p in OPENTRUST_PROJECTS:
    proj_rows.append(p)
append_rows(ws, proj_rows)
print(f"  Added {len(proj_rows)} OpenTrust projects (total: {ws.max_row - 1})")

# 08_DOCS_OFICIAIS
print("Updating 08_DOCS_OFICIAIS...")
ws = wb['08_DOCS_OFICIAIS']
expected_docs = ['doc_id', 'tecnologia', 'caminho', 'uso_principal', 'skills', 'agents', 'projetos',
                 'quando_usar', 'quando_nao_usar', 'consulta_mcp']
existing = [cell.value for cell in ws[1]]
if existing == expected_docs:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_docs}, got={existing}")

doc_rows = []
for d in OPENTRUST_DOCS:
    doc_rows.append(d)
append_rows(ws, doc_rows)
print(f"  Added {len(doc_rows)} OpenTrust docs (total: {ws.max_row - 1})")

# 07_PROMPTS
print("Updating 07_PROMPTS...")
ws = wb['07_PROMPTS']
expected_prompts = ['prompt_id', 'scenario', 'objective', 'agent', 'skill', 'bundle', 'sources',
                    'template', 'expected_output']
existing = [cell.value for cell in ws[1]]
if existing == expected_prompts:
    print("  Headers match ✓")
else:
    print(f"  Headers differ: expected={expected_prompts}, got={existing}")

prompt_rows = []
for p in OPENTRUST_PROMPTS:
    prompt_rows.append(p)
append_rows(ws, prompt_rows)
print(f"  Added {len(prompt_rows)} OpenTrust prompts (total: {ws.max_row - 1})")

# ── Save ──────────────────────────────────────────────────────────────────────

wb.save(XLSX)
print(f"\nSaved to {XLSX}")
